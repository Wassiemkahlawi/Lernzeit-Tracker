import streamlit as st
import pandas as pd
import io
from datetime import datetime, date, timedelta
import uuid
import os
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import plotly.express as px
from pydrive.auth import GoogleAuth
from pydrive.drive import GoogleDrive

# --------------------------------------------
# 1) Page-Config & Sidebar
# --------------------------------------------
st.set_page_config(
    page_title="Lernzeit-Tracker",
    page_icon="📚",
    layout="wide"
)

with st.sidebar:
    st.markdown("## 🧭 Navigation")
    st.markdown("- 🎯 [Zielsystem](#zielsystem-f%C3%BCr-heute)")
    st.markdown("- 📖 [Einträge anzeigen](#bisherige-eintr%C3%A4ge-anzeigen--bearbeiten)")
    st.markdown("- 📅 [Wochenauswertung](#lernzeit-auswertung-dieser-woche)")
    st.markdown("- 🔎 [Filter & Export](#filter--export)")
    st.markdown("- ☁️ [Cloud-Backups](#cloud-backups-google-drive)")
    st.markdown("- 🌡️ [Heatmap](#lernzeit-heatmap-kalender%C3%BCbersicht)")
    st.markdown("- 🗑️ [Datenbank zurücksetzen](#datenbank-zur%C3%BCcksetzen)")
    st.markdown("---")
    st.markdown("### ⏰ Einstellungen")
    # (Falls du Settings brauchst – z.B. Start/Enddatum)

st.markdown("# 📚 Lernzeit-Tracker")
st.markdown("Dein persönliches Tool zum Tracken, Reflektieren und Exportieren deiner Lernzeit.")
st.markdown("---")

# --------------------------------------------
# 2) Cloud-Config & Drive-Funktionen
# --------------------------------------------
CSV_DATEI = "daten.csv"
BACKUP_LOG = "backup_log.txt"
ZIELE_DATEI = "ziele.csv"

# Ordner erstellen falls nicht vorhanden
if not os.path.exists(CSV_DATEI):
    pd.DataFrame(columns=["ID","Fach","Dauer (Minuten)","Datum","Notiz","Tagesziel"]).to_csv(CSV_DATEI, index=False)
if not os.path.exists(ZIELE_DATEI):
    pd.DataFrame(columns=["Datum","Tagesziel"]).to_csv(ZIELE_DATEI, index=False)

@st.cache_resource
def init_drive():
    gauth = GoogleAuth()
    gauth.DEFAULT_SETTINGS['client_config_file'] = "client_secrets.json"
    gauth.LoadCredentialsFile("mycreds.txt")
    if gauth.credentials is None:
        gauth.CommandLineAuth()
    elif gauth.access_token_expired:
        gauth.Refresh()
    else:
        gauth.Authorize()
    gauth.SaveCredentialsFile("mycreds.txt")
    return GoogleDrive(gauth)

def get_or_create_backup_folder(drive, folder_name="Lernzeit_Autobackups"):
    try:
        folder_list = drive.ListFile({
            'q': f"title='{folder_name}' and mimeType='application/vnd.google-apps.folder' and trashed=false"
        }).GetList()
        if folder_list:
            return folder_list[0]['id']
        else:
            folder = drive.CreateFile({
                'title': folder_name,
                'mimeType': 'application/vnd.google-apps.folder'
            })
            folder.Upload()
            return folder['id']
    except Exception as e:
        st.error(f"Fehler beim Erstellen/Abrufen des Backup-Ordners: {e}")
        return None

def upload_to_drive(file_path, file_name="Lernzeit_backup.csv", auto_backup=False):
    try:
        drive = init_drive()
        folder_id = None
        if auto_backup:
            folder_id = get_or_create_backup_folder(drive)
        file_drive = drive.CreateFile({
            'title': file_name,
            'parents': [{'id': folder_id}] if folder_id else []
        })
        file_drive.SetContentFile(file_path)
        file_drive.Upload()
        return file_drive['title']
    except Exception as e:
        st.error(f"Fehler beim Backup: {e}")
        return None

def delete_backup_from_drive(file_id):
    try:
        drive = init_drive()
        file = drive.CreateFile({'id': file_id})
        file.Delete()
        st.success("Backup wurde erfolgreich gelöscht.")
    except Exception as e:
        st.error(f"Fehler beim Löschen des Backups: {e}")

def get_drive_backups_list():
    try:
        drive = init_drive()
        file_list = drive.ListFile({
            'q': "title contains 'Lernzeit_backup' and trashed=false"
        }).GetList()
        backups = []
        for file in file_list:
            if "id" not in file or "title" not in file: 
                continue
            title = file['title']
            erstellt = file.get('createdDate', 'unbekannt')[:10]
            file_id = file.get('id', None)
            typ = "Auto" if "Autobackup" in title else "Manuell"
            if file_id:
                backups.append({
                    "ID": file_id,
                    "Title": title,
                    "Erstellt am": erstellt,
                    "Typ": typ,
                    "Download-Link": f"https://drive.google.com/uc?id={file_id}&export=download"
                })
        return backups
    except Exception as e:
        st.warning(f"Fehler beim Abrufen der Backups: {e}")
        return []

def ist_backup_heute_erfolgt():
    if os.path.exists(BACKUP_LOG):
        with open(BACKUP_LOG, "r") as log:
            letzter_tag = log.read().strip()
            return letzter_tag == str(date.today())
    return False

def backup_log_schreiben():
    with open(BACKUP_LOG, "w") as log:
        log.write(str(date.today()))

# --------------------------------------------
# 3) Manueller & Auto-Backup-Bereich
# --------------------------------------------
st.subheader("☁️ Cloud-Backup (Google Drive)")

# Manueller Backup-Button
if st.button("📤 Backup mit Zeitstempel hochladen", key="manual_backup"):
    try:
        timestamp = datetime.now().strftime("%Y-%m-%d-%H-%M")
        filename = f"Lernzeit_backup_{timestamp}.csv"
        uploaded_name = upload_to_drive(CSV_DATEI, file_name=filename)
        st.success(f"Backup erfolgreich hochgeladen als: {uploaded_name}")
    except Exception as e:
        st.warning(f"Fehler beim Hochladen: {e}")

# Auto-Backup einmal täglich (wenn noch nicht passiert)
if not ist_backup_heute_erfolgt():
    try:
        upload_to_drive(CSV_DATEI, auto_backup=True)
        backup_log_schreiben()
        st.info("🔄 Tägliches Auto-Backup wurde in deinen Drive-Ordner hochgeladen")
    except Exception as e:
        st.warning(f"Auto-Backup fehlgeschlagen: {e}")

st.markdown("---")

# --------------------------------------------
# 4) Drive-Verwaltung mit Filter & Löschfunktion
# --------------------------------------------
st.subheader("☁️ Gespeicherte Cloud-Backups (Drive)")

backups = get_drive_backups_list()
if backups:
    df_backups = pd.DataFrame(backups)
    # Filter Dropdown
    backup_typen = df_backups["Typ"].unique().tolist()
    typ_filter = st.selectbox("Backup-Typ anzeigen", ["Alle"] + backup_typen)

    if typ_filter != "Alle":
        df_backups = df_backups[df_backups["Typ"] == typ_filter]

    df_backups = df_backups.sort_values("Erstellt am", ascending=False).head(5)

    for index, row in df_backups.iterrows():
        col1, col2, col3 = st.columns([4, 2, 1])
        col1.markdown(f"**{row['Title']}**<br>{row['Erstellt am']}", unsafe_allow_html=True)
        col2.markdown(f"[⬇️ Herunterladen]({row['Download-Link']})", unsafe_allow_html=True)
        if col3.button("🗑️ Löschen", key=f"delete_{row['ID']}"):
            delete_backup_from_drive(row["ID"])
            st.rerun()
else:
    st.info("Noch keine Backups gefunden oder Fehler beim Abrufen")

st.markdown("---")

# --------------------------------------------
# 5) Daten & CSV-Initialisierung
# --------------------------------------------
try:
    df = pd.read_csv(CSV_DATEI)
    df["Datum"] = pd.to_datetime(df["Datum"], errors="coerce")
    df = df.dropna(subset=["Datum"])

    if "Notiz" not in df.columns:
        df["Notiz"] = ""
    else:
        df["Notiz"] = df["Notiz"].fillna("")

    if "Tagesziel" not in df.columns:
        df["Tagesziel"] = ""

    # Automatische Bereinigung (älter als 6 Monate)
    df_orig = df.copy()
    sechs_monate_zurueck = datetime.today() - timedelta(days=180)
    df = df[df["Datum"] >= sechs_monate_zurueck]
    geloeschte_anzahl = len(df_orig) - len(df)
    if geloeschte_anzahl > 0:
        st.info(f"✅ {geloeschte_anzahl} Einträge älter als 6 Monate entfernt")

    df.to_csv(CSV_DATEI, index=False)

except Exception:
    df = pd.DataFrame(columns=["ID","Fach","Dauer (Minuten)","Datum","Notiz","Tagesziel"])
    df.to_csv(CSV_DATEI, index=False)
    st.warning("🚨 daten.csv war beschädigt und wurde neu erstellt.")

# --------------------------------------------
# 6) Titel & Eingabemaske
# --------------------------------------------
st.subheader("✏️ Neuen Eintrag hinzufügen")

Fach = st.text_input("📘 Fach (z. B. Mathe, Englisch)")
Dauer = st.number_input("⏱️ Lernzeit in Minuten", min_value=1, step=1)
Datum = st.date_input("📅 Datum", value=datetime.today())
Notiz = st.text_area("🗒️ Notiz (optional)", height=100)

# Tagesziel in der Maske anzeigen (falls separat genutzt)
if "Tagesziel" in df.columns:
    st.info(f"Dein Lernziel für heute: **{int(df[df['Datum'] == pd.to_datetime(date.today())]['Tagesziel'].sum() or 0)} Minuten**")

if st.button("💾 Eintrag speichern"):
    if Fach:
        Eintrag_id = str(uuid.uuid4())
        neuer_eintrag = pd.DataFrame([[Eintrag_id, Fach, Dauer, Datum, Notiz, "90"]],
            columns=["ID","Fach","Dauer (Minuten)","Datum","Notiz","Tagesziel"])

        df_neu = pd.concat([df, neuer_eintrag], ignore_index=True)
        df_neu.to_csv(CSV_DATEI, index=False)
        st.success("✅ Eintrag wurde gespeichert")
        upload_to_drive(CSV_DATEI)  # Automatisches Backup
    else:
        st.warning("ℹ️ Bitte gib ein Fach ein!")

st.markdown("---")

# --------------------------------------------
# 7) Einträge anzeigen & bearbeiten
# --------------------------------------------
with st.expander("📖 Bisherige Einträge anzeigen / bearbeiten"):
    st.markdown("Hier kannst du vorhandene Einträge löschen oder Notizen bearbeiten.")
    if not df.empty:
        df = df.sort_values("Datum", ascending=False)
        heute = datetime.today().date()

        for index, row in df.iterrows():
            col1, col2, col3, col4, col5 = st.columns([2,2,2,3,1])
            ist_heute = (row["Datum"].date() == heute)
            style = "background-color: #28a745; color: white; padding:4px; border-radius:4px;" if ist_heute else ""

            col1.markdown(f'<div style="{style}">{row["Fach"]}</div>', unsafe_allow_html=True)
            col2.markdown(f'<div style="{style}">{row["Dauer (Minuten)"]} Min.</div>', unsafe_allow_html=True)
            col3.markdown(f'<div style="{style}">{row["Datum"].date()}</div>', unsafe_allow_html=True)

            with col4:
                neue_notiz = st.text_input(f"", value=row.get("Notiz",""), key=f"notiz_{row['ID']}")
                if st.button("💾 Speichern", key=f"save_notiz_{row['ID']}"):
                    if neue_notiz.strip():
                        df.loc[df["ID"] == row["ID"], "Notiz"] = neue_notiz
                        df.to_csv(CSV_DATEI, index=False)
                        st.success("✅ Notiz aktualisiert")
                        st.rerun()
                    else:
                        st.warning("ℹ️ Notiz darf nicht leer sein")

            if col5.button("🗑️ Löschen", key=f"delete_entry_{row['ID']}"):
                df = df[df["ID"] != row["ID"]]
                df.to_csv(CSV_DATEI, index=False)
                st.success("🗑️ Eintrag gelöscht")
                st.rerun()
    else:
        st.info("Keine Einträge vorhanden.")

st.markdown("---")

# --------------------------------------------
# 8) Zielsystem für heute
# --------------------------------------------
st.subheader("🎯 Zielsystem für heute")
ziel_minuten = st.number_input("🎯 Tagesziel (Minuten)", min_value=10, step=10, value=90)

# Fortschritt für heute berechnen
heute = datetime.today().date()
schriftliche_df = df.copy()
schriftliche_df["Datum"] = pd.to_datetime(schriftliche_df["Datum"])
heutige_eintraege = schriftliche_df[schriftliche_df["Datum"].dt.date == heute]
gesamtzeit_heute = heutige_eintraege["Dauer (Minuten)"].sum()
fortschritt = gesamtzeit_heute / ziel_minuten if ziel_minuten else 0
prozent = min(100, int(fortschritt * 100))

st.progress(prozent)
st.markdown(f"**{gesamtzeit_heute} / {ziel_minuten} Minuten gelernt ({prozent} %)**")
if fortschritt >= 1:
    st.success("✅ Ziel erreicht – stark, Bruder!")
elif fortschritt >= 0.5:
    st.info("ℹ️ Halbzeit! Du bist gut dabei.")
else:
    st.warning("💪 Weiter so! Du schaffst das.")

# Ziel speichern (erneut) – falls du es weiterhin in daten.csv führst
try:
    df_ziele = pd.read_csv(ZIELE_DATEI)
    df_ziele["Datum"] = pd.to_datetime(df_ziele["Datum"])
    if not (df_ziele["Datum"] == pd.to_datetime(heute)).any():
        pd.concat([df_ziele, pd.DataFrame([[heute, ziel_minuten]], columns=["Datum", "Tagesziel"])], ignore_index=True).to_csv(ZIELE_DATEI, index=False)
except Exception:
    pass

st.markdown("---")

# --------------------------------------------
# 9) Wochenauswertung
# --------------------------------------------
with st.expander("📅 Lernzeit-Auswertung dieser Woche"):
    st.markdown("Diese Tabelle zeigt dir, wie viele Minuten du pro Tag dieser Woche gelernt hast.")
    heute_ts = pd.Timestamp.today().normalize()
    wochenstart = heute_ts - timedelta(days=heute_ts.weekday())
    wochenende = wochenstart + timedelta(days=6)
    if not df.empty and pd.api.types.is_datetime64_any_dtype(df["Datum"]):
        wochen_eintraege = df[(df["Datum"] >= wochenstart) & (df["Datum"] <= wochenende)]
    else:
        wochen_eintraege = pd.DataFrame(columns=df.columns)
        st.info("Keine Lernzeit-Einträge für diese Woche vorhanden.")
    wochen_summary = wochen_eintraege.groupby(df["Datum"].dt.strftime("%A"))["Dauer (Minuten)"].sum()
    TageListe = ["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"]
    wochen_summary = wochen_summary.reindex(TageListe, fill_value=0)
    st.table(wochen_summary.rename_axis("Wochentag").reset_index().rename(columns={"Dauer (Minuten)": "Minuten"}))

st.markdown("---")

# --------------------------------------------
# 10) Filter & Export
# --------------------------------------------
with st.expander("🔎 Filter & Export"):
    st.markdown("### 🧠 Filter")
    # Fach-Filter
    Faecher = df["Fach"].unique().tolist()
    Ausgewaehltes_fach = st.selectbox("Nach Fach filtern", ["Alle"] + Faecher)

    # Datumsbereich-Filter
    if not df.empty:
        start_vorschlag = df["Datum"].min().date()
        end_vorschlag = df["Datum"].max().date()
    else:
        start_vorschlag = date.today()
        end_vorschlag = date.today()

    start_datum = st.date_input("Startdatum", value=start_vorschlag)
    end_datum = st.date_input("Enddatum", value=end_vorschlag)

    gefiltert = df.copy()
    if Ausgewaehltes_fach != "Alle":
        gefiltert = gefiltert[gefiltert["Fach"] == Ausgewaehltes_fach]
    gefiltert = gefiltert[(gefiltert["Datum"] >= pd.to_datetime(start_datum)) & (gefiltert["Datum"] <= pd.to_datetime(end_datum))]

    if not gefiltert.empty:
        st.markdown("### 📤 Export")
        df_export = gefiltert.copy()
        df_export["Datum"] = df_export["Datum"].dt.strftime("%d.%m.%Y")
        wb2 = Workbook()
        ws2 = wb2.active
        ws2.title = "Lernzeit"
        for r in dataframe_to_rows(df_export, index=False, header=True):
            ws2.append(r)
        for col in ws2.columns:
            max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws2.column_dimensions[col_letter].width = max_len + 2
        buf = io.BytesIO()
        wb2.save(buf)
        buf.seek(0)
        st.download_button(
            label="⬇️ Excel-Datei herunterladen",
            data=buf,
            file_name="lernzeit_export.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    else:
        st.info("Keine Daten zum Exportieren.")

st.markdown("---")

# --------------------------------------------
# 11) Lernzeit nach Fach (Diagram)
# --------------------------------------------
with st.expander("📈 Lernzeit pro Fach (Diagram)"):
    st.markdown("Hier siehst du deine gesamte Lernzeit pro Fach.")
    if not gefiltert.empty:
        statistik = gefiltert.groupby("Fach")["Dauer (Minuten)"].sum()
        st.bar_chart(statistik)
    else:
        st.info("Keine Daten für die Visualisierung verfügbar.")

st.markdown("---")

# --------------------------------------------
# 12) Lernzeit-Heatmap
# --------------------------------------------
with st.expander("🌡️ Lernzeit-Heatmap (Kalenderübersicht)"):
    st.markdown("Diese Übersicht zeigt dir, an welchen Tagen du wie viel gelernt hast.")
    if not gefiltert.empty:
        gefiltert["Datum"] = pd.to_datetime(gefiltert["Datum"])
        heatmap_data = gefiltert.groupby(gefiltert["Datum"].dt.date)["Dauer (Minuten)"].sum().reset_index()
        heatmap_data["Datum"] = pd.to_datetime(heatmap_data["Datum"])
        heatmap_data["Wochentag"] = heatmap_data["Datum"].dt.day_name()
        heatmap_data["Kalenderwoche"] = heatmap_data["Datum"].dt.isocalendar().week
        heatmap_data["Wochentag"] = pd.Categorical(
            heatmap_data["Wochentag"],
            categories=["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"],
            ordered=True
        )
        fig = px.density_heatmap(
            data_frame=heatmap_data,
            x="Kalenderwoche",
            y="Wochentag",
            z="Dauer (Minuten)",
            color_continuous_scale="Viridis",
            title="Lernzeit pro Tag (Heatmap)",
            labels={"Dauer (Minuten)": "Lernzeit (Min)"}
        )
        st.plotly_chart(fig, use_container_width=True)
    else:
        st.info("Keine Daten für die Heatmap verfügbar.")

st.markdown("---")

# --------------------------------------------
# 13) Zielverlauf (Tagesziele)
# --------------------------------------------
with st.expander("📆 Zielverlauf"):
    st.markdown("Die Übersicht zeigt dir deine gesetzten Tagesziele.")
    try:
        df_ziele = pd.read_csv(ZIELE_DATEI)
        df_ziele["Datum"] = pd.to_datetime(df_ziele["Datum"])
        st.dataframe(df_ziele.sort_values("Datum", ascending=False))
    except Exception:
        st.info("Keine Zielhistorie vorhanden.")

st.markdown("---")

# --------------------------------------------
# 14) Datenbank zurücksetzen
# --------------------------------------------
with st.expander("🗑️ Datenbank zurücksetzen"):
    st.markdown("Achtung: Dadurch werden alle Einträge endgültig gelöscht!")
    if st.checkbox("Ich bin sicher, alle Einträge zu löschen"):
        if st.button("🔥 Alles löschen", key="reset_all"):
            df_leer = pd.DataFrame(columns=["ID","Fach","Dauer (Minuten)","Datum","Notiz","Tagesziel"])
            df_leer.to_csv(CSV_DATEI, index=False)
            st.success("✅ Alle Einträge wurden gelöscht!")
            st.rerun()
